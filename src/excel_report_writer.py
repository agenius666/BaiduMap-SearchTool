# Copyright 2023 agenius666
# GitHub: https://github.com/agenius666/BaiduMap-SearchTool
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelWriter:
    @staticmethod
    def write(output_path, processed_data, template_path, config, progress_callback=None):
        try:
            template_df = pd.read_excel(template_path, dtype={'分组': str})
            grouped = template_df.groupby('分组')

            # --- 1. 生成字段名映射表 ---
            field_name_map = {}
            enabled_items = [item for item in config["config"]["items"] if item["enabled"]]
            for item in enabled_items:
                original_name = item["name"]
                if original_name == "X米半径范围内公共交通线路数":
                    radius = item.get("radius", "X")
                    new_name = f"{radius}米半径范围内公共交通线路数"
                    field_name_map[original_name] = new_name

            with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                if 'Sheet1' in writer.book.sheetnames:
                    writer.book.remove(writer.book['Sheet1'])

                # 遍历分组生成Sheet
                for group_id, group_data in grouped:
                    sheet_name = f"分组{group_id}"
                    writer.book.create_sheet(title=sheet_name)
                    worksheet = writer.book[sheet_name]

                    # --- 2. 写入表头---
                    headers = ["类目"] + list(group_data['类型'].unique())
                    worksheet.append(headers)

                    # --- 3. 写入数据行---
                    ordered_fields = ["名称"] + [item["name"] for item in enabled_items]
                    for field in ordered_fields:
                        # 生成数据行第一列的值
                        display_field = field_name_map.get(field, field)  
                        row = [display_field]  

                        # 填充数据
                        for col in headers[1:]:
                            community_name = group_data[group_data['类型'] == col]['小区'].values[0]
                            value = processed_data.get(community_name, {}).get(field, "无数据")
                            row.append(str(value))

                        worksheet.append(row)

            return True
        except Exception as e:
            raise RuntimeError(f"Excel生成失败: {str(e)}")
